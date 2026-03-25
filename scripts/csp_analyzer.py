DOUBLE_BOOKING_THRESHOLD_PCT: float = 13.06  # matches the raw dataset's actual double-booking rate

import pandas as pd
from typing import Dict, List, Tuple, Optional, Set
from dataclasses import dataclass, field
from collections import defaultdict
import random
import math

from data_loader import TimetableDataLoader, DATA_RAW
import re


@dataclass
class TimeSlot:
    """Represents a teaching timeslot."""
    day: str
    start_hour: float
    end_hour: float
    
    def overlaps(self, other: 'TimeSlot') -> bool:
        """Check if this timeslot overlaps with another."""
        if self.day != other.day:
            return False
        return not (self.end_hour <= other.start_hour or self.start_hour >= other.end_hour)
    
    def __hash__(self):
        return hash((self.day, self.start_hour, self.end_hour))


@dataclass
class Event:
    """Represents a teaching event."""
    event_id: str
    module_code: str
    duration_minutes: int
    event_size: int
    event_type: str
    is_whole_class: bool
    online: bool = False
    campus: str = 'Unknown'
    weeks: Set[int] = field(default_factory=set)
    assigned_slot: Optional[TimeSlot] = None
    assigned_room: Optional[str] = None
    effective_size: int = 0


@dataclass
class Room:
    """Represents a teaching room."""
    name: str
    capacity: int
    room_type: str
    building: str
    campus: str = 'Unknown'


@dataclass
class CSPResult:
    """Result of a CSP feasibility check."""
    is_feasible: bool
    events_scheduled: int
    events_unscheduled: int
    capacity_violations: int
    clash_count: int           # total double-booked room-slot pairs (informational)
    double_booking_rate: float = 0.0  # % of room-slots that are double-booked (soft constraint)
    double_booking_threshold: float = DOUBLE_BOOKING_THRESHOLD_PCT
    binding_constraints: List[str] = field(default_factory=list)
    soft_warnings: List[str] = field(default_factory=list)


class TimetableCSP:
    """CSP model for timetabling with optimization."""
    
    def __init__(self, scenario: str):
        self.scenario = scenario
        self.loader = TimetableDataLoader()
        self.available_slots: List[TimeSlot] = []
        self.rooms: Dict[str, Room] = {}
        self.events: Dict[str, Event] = {}
        self.room_schedule: Dict[str, List[Tuple[TimeSlot, str]]] = defaultdict(list)  # room -> [(slot, event_id)]
        self.student_schedules: Dict[str, List[TimeSlot]] = defaultdict(list)
        self.campus_hours: Dict[str, Dict[str, Tuple[int, int]]] = {}  # campus -> {day -> (start, end)}
        self.allowed_double_bookings: Dict[str, Set[str]] = defaultdict(set) # event_id -> set of event_ids allowed to overlap
        self.compulsory_conflicts: Dict[str, Set[str]] = defaultdict(set)  # event_id -> set of event_ids that must not overlap
        self.travel_times: Dict[tuple, int] = {}  # (campus_from, campus_to) -> minutes
        
        self._load_campus_constraints()
        self._init_available_slots()
        self._init_rooms()
        self._init_events()
        self._build_compulsory_conflicts()
        self.travel_times = self.loader.travel_times

    
    def _init_available_slots(self):
        """Initialize available timeslots based on scenario."""
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        
        for day in days:
            if self.scenario == 'baseline':
                hours = range(9, 18)  # 9am-6pm
            elif self.scenario == 'scenario_a':
                hours = range(9, 17)  # 9am-5pm
            elif self.scenario == 'scenario_b':
                hours = range(9, 12) if day == 'Friday' else range(9, 18)
            else:
                hours = range(9, 18)
            
            for hour in hours:
                self.available_slots.append(TimeSlot(day, hour, hour + 1))
    
    def _load_campus_constraints(self):
        """Load campus-specific teaching hours from Room Constraints sheet."""
        try:
            constraints_df = pd.read_excel(DATA_RAW / "Rooms_and_Room_Types.xlsx", sheet_name='Room Constraints')
            days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
            
            for _, row in constraints_df.iterrows():
                campus = row.get('Campus')
                if pd.isna(campus):
                    continue
                campus = str(campus).strip()
                
                if campus not in self.campus_hours:
                    self.campus_hours[campus] = {}
                
                for day in days:
                    hours_str = row.get(day) if day in constraints_df.columns else row.get('Teaching Hours')
                    if pd.isna(hours_str):
                        # Default to 9-6
                        self.campus_hours[campus][day] = (9, 18)
                    else:
                        # Parse "9am-6pm" or "9am-1pm" format
                        match = re.search(r'(\d+)(?:am)?\s*-\s*(\d+)(?:pm)?', str(hours_str))
                        if match:
                            start = int(match.group(1))
                            end = int(match.group(2))
                            if end < start:  # Handle pm conversion
                                end += 12
                            self.campus_hours[campus][day] = (start, end)
                        else:
                            self.campus_hours[campus][day] = (9, 18)
        except Exception as e:
            print(f"Warning: Could not load campus constraints: {e}")
            # Fallback to empty
            self.campus_hours = {}
    
    def _init_rooms(self):
        """Initialize room data."""
        rooms_df = self.loader.rooms
        for _, row in rooms_df.iterrows():
            room = Room(
                name=str(row['Id']),
                capacity=int(row['Capacity']) if pd.notna(row['Capacity']) else 0,
                room_type=str(row.get('Room Type', 'Unknown')),
                building=str(row.get('Building', 'Unknown')),
                campus=str(row.get('Campus', 'Unknown')) if pd.notna(row.get('Campus')) else 'Unknown'
            )
            self.rooms[room.name] = room
    
    def _parse_weeks(self, weeks_str: str) -> Set[int]:
        """Convert a weeks string like '1-5, 7-10' or '1' into a set of integers."""
        if pd.isna(weeks_str):
            return {1} # Default to week 1 to be conservative
            
        weeks = set()
        parts = str(weeks_str).split(',')
        for part in parts:
            part = part.strip()
            if not part: continue
            
            if '-' in part:
                try:
                    start, end = part.split('-')
                    weeks.update(range(int(start), int(end) + 1))
                except ValueError:
                    pass
            else:
                try:
                    weeks.add(int(part))
                except ValueError:
                    pass
        return weeks if weeks else {1}
        
    def _init_events(self):
        """Initialize events from data and extract allowed double bookings."""
        events_df = self.loader.events
        
        # Pre-scan for existing double bookings in the raw data
        # We group by (Day, Start Hour, Room) and any events sharing this are allowed to overlap
        raw_schedule_groups = defaultdict(list)
        for _, row in events_df.iterrows():
            if pd.notna(row['Day']) and pd.notna(row['Start Hour']) and pd.notna(row['Room']):
                key = (row['Day'], row['Start Hour'], row['Room'])
                raw_schedule_groups[key].append(str(row['Event ID']))
                
        for key, event_ids in raw_schedule_groups.items():
            if len(event_ids) > 1:
                # These events all share the same room at the same time in the raw data
                for eid in event_ids:
                    # Allow it to overlap with all others in this group
                    self.allowed_double_bookings[eid].update(e for e in event_ids if e != eid)
        
        for _, row in events_df.iterrows():
            event_id = str(row['Event ID'])
            event = Event(
                event_id=str(row['Event ID']),
                module_code=str(row['Module Code']) if pd.notna(row['Module Code']) else 'Unknown',
                duration_minutes=int(row['Duration (minutes)']) if pd.notna(row['Duration (minutes)']) else 60,
                event_size=int(row['Event Size']) if pd.notna(row['Event Size']) else 0,
                event_type=str(row['Event Type']) if pd.notna(row['Event Type']) else 'Unknown',
                is_whole_class=bool(row['WholeClass']) if 'WholeClass' in row and pd.notna(row['WholeClass']) else False,
                online=False,
                campus=str(row.get('Campus', 'Unknown')) if pd.notna(row.get('Campus')) else 'Unknown',
                weeks=self._parse_weeks(row.get('Weeks'))
            )
            
            original_room_name = str(row['Room']) if pd.notna(row['Room']) else None
            
            # Calculate effective size: event size capped by its original physical room capacity
            effective_size = event.event_size
            if original_room_name and original_room_name in self.rooms:
                orig_capacity = self.rooms[original_room_name].capacity
                effective_size = min(event.event_size, orig_capacity)
            event.effective_size = effective_size
            
            # Set current assignment if within scenario bounds
            if pd.notna(row['Day']) and pd.notna(row['Start Hour']):
                slot = TimeSlot(row['Day'], row['Start Hour'], row['End Hour'])
                
                # Events that don't need a physical room are always kept in place
                room_type = str(row.get('Room Type 1', '')) if pd.notna(row.get('Room Type 1')) else ''
                online_val = row.get('Online Delivery')
                no_room_needed = (room_type == 'No room required') or (pd.notna(online_val) and bool(online_val))
                
                if no_room_needed or self._is_slot_in_bounds(slot):
                    event.assigned_slot = slot
                    event.assigned_room = str(row['Room']) if pd.notna(row['Room']) else None
                    if event.assigned_room:
                        self.room_schedule[event.assigned_room].append((slot, event.event_id))
            
            # Parse Weeks
            weeks_str = row.get('Weeks')
            weeks = set()
            if pd.notna(weeks_str):
                # Handle "1-5, 7-10" format
                parts = str(weeks_str).split(',')
                for part in parts:
                    part = part.strip()
                    if '-' in part:
                        try:
                            start, end = part.split('-')
                            weeks.update(range(int(start), int(end) + 1))
                        except ValueError:
                            pass
                    else:
                        try:
                            weeks.add(int(part))
                        except ValueError:
                            pass
            if not weeks:
                weeks = {1}  # Default to week 1 if missing

            event.weeks = weeks
            
            # Parse Online Delivery
            online_val = row.get('Online Delivery')
            event.online = bool(online_val) if pd.notna(online_val) else False
            
            # Parse Campus
            campus_val = row.get('Campus')
            event.campus = str(campus_val).strip() if pd.notna(campus_val) else 'Unknown'
            
            self.events[event_id] = event
    
    def _build_compulsory_conflicts(self):
        """Build compulsory conflict graph from DPT data.
        
        For each programme-year, identify all compulsory courses and their
        whole-class events. Events from different compulsory courses of the
        same programme-year must not overlap (H4).
        """
        try:
            from compulsory_clash_detector import CompulsoryClashDetector
            whitelist = CompulsoryClashDetector.get_whitelisted_clashes()
        except Exception as e:
            print(f"Warning: Could not load whitelist: {e}")
            whitelist = set()
            
        try:
            dpt = pd.read_excel(DATA_RAW / "2024-5_DPT_Data.xlsx")
        except Exception as e:
            print(f"Warning: Could not load DPT data for conflict graph: {e}")
            return
        
        # Step 1: Get compulsory courses per programme-year
        compulsory = dpt[dpt['Compulsory/Optional'] == 'Compulsory']
        prog_courses: Dict[str, Set[str]] = defaultdict(set)
        for _, row in compulsory.iterrows():
            prog_code = row['Programme Code']
            prog_year = row.get('Programme Year', row.get('ProgYear', ''))
            key = f"{prog_code}_{prog_year}"
            course_code = str(row['Course Code'])
            prog_courses[key].add(course_code)
        
        # Step 2: Map course codes to whole-class event IDs
        events_df = self.loader.events
        whole_class_mask = (events_df['WholeClass'] == True) | (events_df['Event Type'] == 'Lecture')
        wc_events = events_df[whole_class_mask]
        
        course_event_ids: Dict[str, List[str]] = defaultdict(list)
        for _, row in wc_events.iterrows():
            module_code = str(row['Module Code']) if pd.notna(row['Module Code']) else ''
            base_course = module_code.split('_')[0] if module_code else ''
            event_id = str(row['Event ID'])
            if base_course and event_id in self.events:
                course_event_ids[base_course].append(event_id)
        
        # Step 3: Build pairwise conflicts across different compulsory courses
        for prog_key, courses in prog_courses.items():
            courses_list = list(courses)
            for i, c1 in enumerate(courses_list):
                eids1 = course_event_ids.get(c1, [])
                for c2 in courses_list[i + 1:]:
                    eids2 = course_event_ids.get(c2, [])
                    # All events from c1 conflict with all events from c2
                    for eid1 in eids1:
                        for eid2 in eids2:
                            if frozenset([eid1, eid2]) not in whitelist:
                                self.compulsory_conflicts[eid1].add(eid2)
                                self.compulsory_conflicts[eid2].add(eid1)
    
    def _check_compulsory_clash(self, event: Event, slot: TimeSlot) -> bool:
        """Return True if placing event at slot would create a compulsory clash (H4)."""
        conflict_ids = self.compulsory_conflicts.get(event.event_id, set())
        if not conflict_ids:
            return False
        
        for cid in conflict_ids:
            conflict_evt = self.events.get(cid)
            if conflict_evt and conflict_evt.assigned_slot:
                if slot.overlaps(conflict_evt.assigned_slot):
                    if not event.weeks.isdisjoint(conflict_evt.weeks):
                        return True  # H4 violation
        return False
    
    def _is_slot_in_bounds(self, slot: TimeSlot) -> bool:
        """checking if a timeslot is within scenario bounds."""
        # Weekend events are intentional — leave them untouched
        if slot.day in ['Saturday', 'Sunday']:
            return True
        
        # New exemption for intentional midnight/asynchronous placeholders
        if slot.start_hour in [0.0, 0.5]:
            return True
            
        if self.scenario == 'baseline':
            return 9 <= slot.start_hour and slot.end_hour <= 18
        elif self.scenario == 'scenario_a':
            return 9 <= slot.start_hour and slot.end_hour <= 17
        elif self.scenario == 'scenario_b':
            if slot.day == 'Friday':
                return 9 <= slot.start_hour and slot.end_hour <= 12
            return 9 <= slot.start_hour and slot.end_hour <= 18
        return False
    
    def get_displaced_events(self) -> List[Event]:
        """getting events that need to be rescheduled."""
        return [e for e in self.events.values() if e.assigned_slot is None]
    
    def get_scheduled_events(self) -> List[Event]:
        """getting events that are currently scheduled."""
        return [e for e in self.events.values() if e.assigned_slot is not None]
    
    def check_hard_constraints(self) -> CSPResult:
        """
        checking hard and soft constraints and returning feasibility result.

        hard constraints (determine is_feasible):
          - room capacity violations
          - unscheduled events

        soft constraints (tracked separately):
          - room double-bookings: intentional at Edinburgh (shared studios,
            lab pods, Medicine LT block-bookings). Only flagged as a warning
            when the rate exceeds DOUBLE_BOOKING_THRESHOLD_PCT.
        """
        capacity_violations = 0
        clashes = 0
        binding = []
        soft_warnings = []

        # building room schedules
        room_usage = defaultdict(list)  # room -> [(slot, event)]
        total_room_slots = 0

        for event in self.get_scheduled_events():
            # skipping online events from room checks
            if event.online:
                continue

            if event.assigned_room and event.assigned_slot:
                room_usage[event.assigned_room].append((event.assigned_slot, event))
                total_room_slots += 1

        # Hard constraint 1: Capacity
        for event in self.get_scheduled_events():
            if event.assigned_room and event.assigned_room in self.rooms:
                room = self.rooms[event.assigned_room]
                if room.capacity < event.effective_size:
                    capacity_violations += 1

        # soft: count double-booked room-slot pairs
        for room_name, bookings in room_usage.items():
            for i, (slot1, evt1) in enumerate(bookings):
                for slot2, evt2 in bookings[i + 1:]:
                    if slot1.overlaps(slot2):
                        if not evt1.weeks.isdisjoint(evt2.weeks):
                            clashes += 1

        # calculate double-booking rate (as % of total room-slot assignments)
        double_booking_rate = (
            (clashes / total_room_slots * 100) if total_room_slots > 0 else 0.0
        )

        unscheduled = len(self.get_displaced_events())
        scheduled = len(self.get_scheduled_events())

        # feasibility is determined by hard constraints only
        is_feasible = (capacity_violations == 0 and unscheduled == 0)

        if capacity_violations > 0:
            binding.append(f"Room capacity violations: {capacity_violations}")
        if unscheduled > 0:
            binding.append(f"Unscheduled events: {unscheduled}")

        # double bookings: soft warning only if above threshold
        if double_booking_rate > DOUBLE_BOOKING_THRESHOLD_PCT:
            soft_warnings.append(
                f"Double-booking rate {double_booking_rate:.1f}% exceeds "
                f"threshold ({DOUBLE_BOOKING_THRESHOLD_PCT}%): {clashes} overlapping room-slot pairs"
            )
        else:
            soft_warnings.append(
                f"Double-booking rate {double_booking_rate:.1f}% is within "
                f"acceptable threshold ({DOUBLE_BOOKING_THRESHOLD_PCT}%): {clashes} pairs"
            )

        return CSPResult(
            is_feasible=is_feasible,
            events_scheduled=scheduled,
            events_unscheduled=unscheduled,
            capacity_violations=capacity_violations,
            clash_count=clashes,
            double_booking_rate=round(double_booking_rate, 2),
            double_booking_threshold=DOUBLE_BOOKING_THRESHOLD_PCT,
            binding_constraints=binding,
            soft_warnings=soft_warnings,
        )
    
    def find_available_slot(self, event: Event) -> Optional[Tuple[TimeSlot, str]]:
        """finding an available slot and room for an event using greedy search."""
        duration_hours = event.duration_minutes / 60
        
        for slot in self.available_slots:
            # create extended slot for event duration
            extended_slot = TimeSlot(slot.day, slot.start_hour, slot.start_hour + duration_hours)
            
            if not self._is_slot_in_bounds(extended_slot):
                continue
            
            # Hard constraint 4: Compulsory clash avoidance
            if self._check_compulsory_clash(event, extended_slot):
                continue
            
            # find suitable room (must match campus)
            for room_name, room in self.rooms.items():
                # Hard constraint 1: Capacity
                if room.capacity < event.effective_size:
                    continue
                
                # campus match
                if event.campus != 'Unknown' and room.campus != 'Unknown':
                    if event.campus != room.campus:
                        continue
                
                # Check if room is available
                is_available = True
                for (booked_slot, booked_event_id) in self.room_schedule.get(room_name, []):
                    if extended_slot.overlaps(booked_slot):
                        # Is it an explicitly allowed double booking?
                        if booked_event_id in self.allowed_double_bookings.get(event.event_id, set()):
                            continue # Valid overlap
                            
                        booked_event = self.events[booked_event_id]
                        if not event.weeks.isdisjoint(booked_event.weeks):
                            is_available = False
                            break
                
                if is_available:
                    return (extended_slot, room_name)
        
        return None
    
    def greedy_reschedule(self) -> int:
        """attempt to reschedule displaced events using greedy approach."""
        displaced = self.get_displaced_events()
        scheduled_count = 0
        
        # sort by size (larger events first - harder to place)
        displaced.sort(key=lambda e: -e.event_size)
        
        for event in displaced:
            result = self.find_available_slot(event)
            if result:
                slot, room = result
                event.assigned_slot = slot
                event.assigned_room = room
                self.room_schedule[room].append((slot, event.event_id))
                scheduled_count += 1
        
        return scheduled_count
    
    # ------------------------------------------------------------------
    # Schedule management helpers
    # ------------------------------------------------------------------

    def _assign_event(self, event: Event, slot: TimeSlot, room: str) -> None:
        """assigning an event to a slot and room, updating the schedule index."""
        event.assigned_slot = slot
        event.assigned_room = room
        self.room_schedule[room].append((slot, event.event_id))

    def _unassign_event(self, event: Event) -> None:
        """unassigning an event from its current slot and room."""
        if event.assigned_room:
            # Rebuild list without this event
            self.room_schedule[event.assigned_room] = [
                item for item in self.room_schedule[event.assigned_room] 
                if item[1] != event.event_id
            ]
        event.assigned_slot = None
        event.assigned_room = None

    def _is_slot_room_available(self, event: Event, slot: TimeSlot,
                                 room_name: str) -> bool:
        """returning True if event can be placed in slot+room (ignores event's own booking)."""
        if not self._is_slot_in_bounds(slot):
            return False
        
        # Hard constraint 4: Compulsory clash avoidance
        if self._check_compulsory_clash(event, slot):
            return False
            
        room = self.rooms.get(room_name)
        if not room or room.capacity < event.effective_size:
            return False
        if event.campus != 'Unknown' and room.campus != 'Unknown':
            if event.campus != room.campus:
                return False
        for booked_slot, booked_event_id in self.room_schedule.get(room_name, []):
            if booked_event_id == event.event_id:
                continue   # skip the event's own booking if any
            if slot.overlaps(booked_slot):
                # Is it an explicitly allowed double booking?
                if booked_event_id in self.allowed_double_bookings.get(event.event_id, set()):
                    continue # Valid overlap
                    
                booked_event = self.events.get(booked_event_id)
                if booked_event and not event.weeks.isdisjoint(booked_event.weeks):
                    return False
        return True

    # Simulated Annealing

    def simulated_annealing(self, max_iterations: int = 10_000,
                             initial_temp: float = 10.0,
                             cooling_rate: float = 0.998) -> int:
        """
        simulated annealing optimizer for re-inserting displaced events.

        two move types per iteration:

        place  (40%): try to insert a random unscheduled event into a
                        truly free slot (same as greedy but randomised).

        evict  (60%): pick a scheduled event S whose room/slot an
                        unscheduled event U could occupy.  Evict S,
                        place U, then try to re-home S elsewhere.
                        - If S re-homes: net improvement (delta E = -1) -> always accept
                        - If S can't re-home: neutral (delta E = 0, different event
                          is now unscheduled) -> accept with Metropolis criterion
                        This breaks greedy local optima by letting easier events
                        be temporarily displaced to make room for harder ones.

        acceptance rule (Metropolis criterion):
          delta E <= 0  -> always accept
          delta E > 0  -> accept with probability  P = exp(-delta E / T)

        returns the number of additional events scheduled beyond greedy.
        """
        displaced = self.get_displaced_events()
        if not displaced:
            return 0

        initial_unscheduled = len(displaced)
        temp = initial_temp
        scheduled_events = self.get_scheduled_events()

        for _ in range(max_iterations):
            if not displaced:
                break

            # move: direct placement (40%)
            if random.random() < 0.4 or not scheduled_events:
                event = random.choice(displaced)
                result = self.find_available_slot(event)
                if result:
                    slot, room = result
                    self._assign_event(event, slot, room)
                    displaced.remove(event)
                    scheduled_events.append(event)

            # move: evict-and-swap (60%)
            else:
                event_u = random.choice(displaced)

                # candidates: scheduled events whose room is large enough
                # and on the right campus for event_u
                candidates = [
                    evt for evt in scheduled_events
                    if (not evt.online
                        and evt.assigned_room
                        and evt.assigned_slot
                        and (room := self.rooms.get(evt.assigned_room)) is not None
                        and room.capacity >= event_u.effective_size
                        and (event_u.campus == 'Unknown'
                             or room.campus == 'Unknown'
                             or event_u.campus == room.campus))
                ]
                if not candidates:
                    temp *= cooling_rate
                    continue

                event_s = random.choice(candidates)
                saved_slot = event_s.assigned_slot
                saved_room = event_s.assigned_room

                # evicting S temporarily
                self._unassign_event(event_s)

                # checking if U fits in S's old slot/room
                # IMPORTANT: U might have a different duration than S, so we must 
                # recreate the TimeSlot with U's duration at S's start time.
                u_duration_hours = event_u.duration_minutes / 60
                u_slot = TimeSlot(saved_slot.day, saved_slot.start_hour, saved_slot.start_hour + u_duration_hours)
                
                if self._is_slot_room_available(event_u, u_slot, saved_room):
                    self._assign_event(event_u, u_slot, saved_room)
                    displaced.remove(event_u)
                    scheduled_events.append(event_u)

                    # trying to re-home S
                    result_s = self.find_available_slot(event_s)
                    if result_s:
                        # S re-homed → net improvement (delta E = -1), always accept
                        slot_s, room_s = result_s
                        self._assign_event(event_s, slot_s, room_s)
                    else:
                        # S homeless → neutral swap (delta E = 0): U placed, S now displaced.
                        # Accept with Metropolis probability (helps exploration).
                        delta_e = 1
                        if random.random() < math.exp(-delta_e / max(temp, 1e-9)):
                            displaced.append(event_s)
                            scheduled_events.remove(event_s)
                        else:
                            # Reject: undo the whole move
                            self._unassign_event(event_u)
                            displaced.append(event_u)
                            scheduled_events.remove(event_u)
                            self._assign_event(event_s, saved_slot, saved_room)
                else:
                    # U can't fit in S's slot/room → restore S
                    self._assign_event(event_s, saved_slot, saved_room)

            temp *= cooling_rate

        return initial_unscheduled - len(displaced)

    # Timetable export

    def export_timetable(self, output_path: str) -> str:
        """
        exporting the best-effort timetable to an Excel file.

        two sheets are produced:
          'Timetable'   – Every event that was successfully scheduled,
                          showing its final slot and room.  Events that
                          were re-assigned by the optimiser are flagged
                          so it's clear what changed.
          'Violations'  – Events that could NOT be placed within the
                          scenario's allowed hours, along with a plain-
                          English reason so the university knows exactly
                          what still needs a manual solution.
        """
        from pathlib import Path
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        # Collect original timeslots from the raw events data
        raw_events = self.loader.events.copy()
        orig_slots = {}  # event_id -> (original_day, original_start)
        for _, row in raw_events.iterrows():
            orig_slots[str(row['Event ID'])] = {
                'orig_day':   row.get('Day'),
                'orig_start': row.get('Start Hour'),
                'orig_room':  row.get('Room'),
                'module':     row.get('Module Code', ''),
                'event_type': row.get('Event Type', ''),
                'event_size': row.get('Event Size', 0),
                'duration':   row.get('Duration (minutes)', 0),
                'campus':     row.get('Campus', ''),
                'weeks':      row.get('Weeks', ''),
            }

        # building scheduled sheet rows
        def format_time(hour_float):
            if pd.isna(hour_float) or hour_float == '':
                return ''
            try:
                h = int(hour_float)
                m = int(round((float(hour_float) - h) * 60))
                if m == 60:
                    h += 1
                    m = 0
                return f"{h:02d}:{m:02d}"
            except (ValueError, TypeError):
                return ''

        scheduled_rows = []
        for event in self.get_scheduled_events():
            orig = orig_slots.get(event.event_id, {})
            orig_day   = orig.get('orig_day', '')
            orig_start = orig.get('orig_start', '')
            new_day    = event.assigned_slot.day   if event.assigned_slot else ''
            new_start  = event.assigned_slot.start_hour if event.assigned_slot else ''
            new_end    = event.assigned_slot.end_hour   if event.assigned_slot else ''

            # was this event re-assigned by the optimiser?
            was_moved = (orig_day != new_day or orig_start != new_start)

            # capacity violation flag - checking against EFFECTIVE size, which is what the solver uses
            room_obj = self.rooms.get(event.assigned_room or '')
            cap_ok = (room_obj is None or event.effective_size <= room_obj.capacity)

            scheduled_rows.append({
                'Event ID':         event.event_id,
                'Module Code':      orig.get('module', ''),
                'Event Type':       orig.get('event_type', ''),
                'Event Size':       event.event_size,
                'Effective Size':   event.effective_size,
                'Duration (min)':   orig.get('duration', event.duration_minutes),
                'Campus':           orig.get('campus', event.campus),
                'Weeks':            orig.get('weeks', ''),
                'Original Day':     orig_day,
                'Original Start':   format_time(orig_start),
                'New Day':          new_day,
                'New Start':        format_time(new_start),
                'New End':          format_time(new_end),
                'Room':             event.assigned_room or '',
                'Room Capacity':    room_obj.capacity if room_obj else '',
                'Re-assigned':      'Yes' if was_moved else 'No',
                'Capacity OK':      'Yes' if cap_ok else 'Violation',
            })

        # building violations sheet rows
        violation_rows = []
        rooms_df = self.loader.rooms.set_index('Description') if len(self.loader.rooms) else None

        for event in self.get_displaced_events():
            orig = orig_slots.get(event.event_id, {})

            # diagnosing why it couldn't be placed
            reasons = []
            # check if any room is big enough at all
            big_enough = [r for r in self.rooms.values() if r.capacity >= event.event_size]
            if not big_enough:
                reasons.append(f"No room fits event size {event.event_size}")
            else:
                # Campus constraint
                campus_rooms = [r for r in big_enough
                                if event.campus in ('Unknown', r.campus, '')
                                or r.campus == 'Unknown']
                if not campus_rooms:
                    reasons.append(f"No room on campus '{event.campus}' with capacity >= {event.event_size}")
                else:
                    reasons.append("All suitable slots occupied after optimisation")

            violation_rows.append({
                'Event ID':      event.event_id,
                'Module Code':   orig.get('module', ''),
                'Event Type':    orig.get('event_type', ''),
                'Event Size':    event.event_size,
                'Effective Size': event.effective_size,
                'Duration (min)': orig.get('duration', event.duration_minutes),
                'Campus':        orig.get('campus', event.campus),
                'Weeks':         orig.get('weeks', ''),
                'Original Day':  orig.get('orig_day', ''),
                'Original Start': format_time(orig.get('orig_start', '')),
                'Reason':        '; '.join(reasons),
                'Suggested Action': 'Manual scheduling required',
            })

        # writing to Excel
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()

        # Sheet 1: Timetable
        ws1 = wb.active
        ws1.title = 'Timetable'

        header_fill  = PatternFill('solid', fgColor='1F4E79')
        moved_fill   = PatternFill('solid', fgColor='FFF2CC')   # light yellow
        cap_vio_fill = PatternFill('solid', fgColor='FFE0E0')   # light red
        header_font  = Font(bold=True, color='FFFFFF')

        if scheduled_rows:
            cols = list(scheduled_rows[0].keys())
            for col_idx, col in enumerate(cols, 1):
                cell = ws1.cell(row=1, column=col_idx, value=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for row_idx, row_data in enumerate(scheduled_rows, 2):
                for col_idx, col in enumerate(cols, 1):
                    ws1.cell(row=row_idx, column=col_idx, value=row_data[col])
                # Highlight re-assigned rows
                if row_data['Re-assigned'] == 'Yes':
                    for col_idx in range(1, len(cols) + 1):
                        ws1.cell(row=row_idx, column=col_idx).fill = moved_fill
                # Highlight capacity violations
                if row_data['Capacity OK'] != 'Yes':
                    ws1.cell(row=row_idx, column=cols.index('Capacity OK') + 1).fill = cap_vio_fill

            # Auto-fit columns
            for col_idx in range(1, len(cols) + 1):
                ws1.column_dimensions[get_column_letter(col_idx)].width = 18

        # Sheet 2: Violations
        ws2 = wb.create_sheet('Violations')
        vio_header_fill = PatternFill('solid', fgColor='8B0000')

        if violation_rows:
            cols2 = list(violation_rows[0].keys())
            for col_idx, col in enumerate(cols2, 1):
                cell = ws2.cell(row=1, column=col_idx, value=col)
                cell.fill = vio_header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for row_idx, row_data in enumerate(violation_rows, 2):
                for col_idx, col in enumerate(cols2, 1):
                    ws2.cell(row=row_idx, column=col_idx, value=row_data[col])

            for col_idx in range(1, len(cols2) + 1):
                ws2.column_dimensions[get_column_letter(col_idx)].width = 22

        ws2.cell(row=1 if not violation_rows else len(violation_rows) + 3,
                 column=1,
                 value=f"Total violations: {len(violation_rows)}")

        wb.save(output_path)
        return output_path


def analyze_scenario(scenario: str, export: bool = True) -> Dict:
    """
    Run full CSP analysis for a scenario.

    If export=True, writes a best-effort timetable Excel file to
    outputs/timetable_<scenario>.xlsx — always produced regardless of
    feasibility, with a Violations sheet for anything that couldn't be placed.
    """
    print(f"\nAnalyzing {scenario}...")

    csp = TimetableCSP(scenario)

    # initial state
    initial_displaced = len(csp.get_displaced_events())
    initial_scheduled = len(csp.get_scheduled_events())

    # check constraints before optimization
    csp.check_hard_constraints()

    # try to reschedule displaced events
    greedy_scheduled = csp.greedy_reschedule()
    sa_scheduled = csp.simulated_annealing()

    # final check
    final_result = csp.check_hard_constraints()

    # exporting timetable (always - best effort)
    export_path = None
    if export:
        from pathlib import Path
        output_dir = Path(__file__).parent.parent / 'outputs'
        output_dir.mkdir(exist_ok=True)
        export_path = str(output_dir / f'timetable_{scenario}.xlsx')
        csp.export_timetable(export_path)
        print(f"  Timetable exported -> {export_path}")

    return {
        'scenario': scenario,
        'initial_scheduled': initial_scheduled,
        'initial_displaced': initial_displaced,
        'greedy_rescheduled': greedy_scheduled,
        'sa_rescheduled': sa_scheduled,
        'final_scheduled': final_result.events_scheduled,
        'final_unscheduled': final_result.events_unscheduled,
        'is_feasible': final_result.is_feasible,
        'capacity_violations': final_result.capacity_violations,
        'clash_count': final_result.clash_count,
        'double_booking_rate': final_result.double_booking_rate,
        'double_booking_threshold': final_result.double_booking_threshold,
        'binding_constraints': final_result.binding_constraints,
        'soft_warnings': final_result.soft_warnings,
        'export_path': export_path,
    }


if __name__ == "__main__":
    for scenario in ['baseline', 'scenario_a', 'scenario_b']:
        result = analyze_scenario(scenario, export=True)
        print(f"\n{'='*50}")
        print(f"Scenario: {scenario}")
        print(f"Initial: {result['initial_scheduled']} scheduled, {result['initial_displaced']} displaced")
        print(f"After optimization: {result['final_scheduled']} scheduled, {result['final_unscheduled']} unscheduled")
        if result['binding_constraints']:
            print(f"Hard constraint violations: {result['binding_constraints']}")
        if result['export_path']:
            print(f"Timetable saved: {result['export_path']}")
